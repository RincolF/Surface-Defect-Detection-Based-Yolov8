import sys
import os
import shutil
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QPushButton, QLabel, QFrame, QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox, QFileDialog)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QSize, QTimer
from PyQt5.QtGui import QPixmap, QImage, QIcon
import cv2
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from datetime import datetime
from ultralytics import YOLO
import time
from openpyxl import Workbook

class ImageProcessor:
    def __init__(self, work_dir):
        # 加载模型
        self.model1 = YOLO("./utils/inside_quexian_train3.pt")
        self.model2 = YOLO("./utils/outside_quexian_train5.pt")
        self.model3 = YOLO("./utils/lock_cut_train2.pt")
        
        # 设置文件夹路径
        self.work_dir = work_dir
        self.predict_folder = os.path.join(work_dir, "Predict")
        self.inside_folder = os.path.join(self.predict_folder, "Inside")
        self.outside_folder = os.path.join(self.predict_folder, "Outside")
        
        # 拷贝classes文件
        self.source_folder = './utils'
        self.file1_name = "classes_outside.txt"
        self.file2_name = "classes_inside.txt"
        
    def get_class_name(self, label_folder, label_id):
        classes_file = os.path.join(label_folder, "classes.txt")
        if not os.path.exists(classes_file):
            return "Unknown"
        with open(classes_file, "r") as f:
            classes = [line.strip() for line in f.readlines()]
        return classes[int(label_id)] if int(label_id) < len(classes) else "Unknown"

    def process_folder_for_excel(self, image_folder, label_folder, excel_row, ws):
        image_files = [f for f in os.listdir(image_folder) if f.endswith(".jpg")]
        label_files = [f for f in os.listdir(label_folder) if f.endswith(".txt")]

        for image_file in image_files:
            image_name = os.path.splitext(image_file)[0]
            label_file = f"{image_name}.txt"
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 获取当前时间
            
            if label_file in label_files:
                with open(os.path.join(label_folder, label_file), "r") as f:
                    lines = f.readlines()
                    # 使用集合去重
                    labels = {self.get_class_name(label_folder, line.split()[0]) for line in lines}
                    label_str = ", ".join(labels) if labels else "OK"
            else:
                label_str = "OK"

            # 如果图片已存在于Excel中，则更新标签，否则添加新行
            if image_name not in excel_row:  # 检查是否已存在该图片文件名
                ws.append([image_file, label_str, current_time])  # 添加新行，包含时间信息
                excel_row[image_name] = ws.max_row  # 更新字典，记录行号
            else:
                current_label = ws.cell(row=excel_row[image_name], column=2).value
                if current_label == "OK" and label_str != "OK":
                    # 如果当前标签是OK，但新标签不是OK，则替换OK为新标签
                    ws.cell(row=excel_row[image_name], column=2, value=label_str)
                    ws.cell(row=excel_row[image_name], column=3, value=current_time)  # 更新时间
                elif current_label != "OK" and label_str != "OK":
                    # 如果当前标签和新标签都不是OK，则合并去重
                    combined_labels = set(current_label.split(", ")) | set(label_str.split(", "))
                    ws.cell(row=excel_row[image_name], column=2, value=", ".join(combined_labels))
                    ws.cell(row=excel_row[image_name], column=3, value=current_time)  # 更新时间

    def generate_excel(self):
        # 拷贝classes文件
        source1_file_path = os.path.join(self.source_folder, self.file1_name)
        source2_file_path = os.path.join(self.source_folder, self.file2_name)
        destination1_file_path = os.path.join(self.outside_folder, "labels/classes.txt")
        destination2_file_path = os.path.join(self.inside_folder, "labels/classes.txt")
        
        os.makedirs(os.path.dirname(destination1_file_path), exist_ok=True)
        os.makedirs(os.path.dirname(destination2_file_path), exist_ok=True)
        
        shutil.copy(source1_file_path, destination1_file_path)
        shutil.copy(source2_file_path, destination2_file_path)

        # 创建Excel工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Labels"
        ws.append(["Image Filename", "Labels", "Processing Time"])  # 添加时间列的表头

        excel_row = {}  # 用于跟踪图片文件名对应的Excel行号

        # 处理Inside文件夹
        inside_label_folder = os.path.join(self.inside_folder, "labels")
        self.process_folder_for_excel(self.inside_folder, inside_label_folder, excel_row, ws)

        # 处理Outside文件夹
        outside_label_folder = os.path.join(self.outside_folder, "labels")
        self.process_folder_for_excel(self.outside_folder, outside_label_folder, excel_row, ws)

        # 调整列宽以适应内容
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # 保存Excel文件到工作目录
        excel_path = os.path.join(self.work_dir, "Predict.xlsx")
        wb.save(excel_path)
        print(f"Result has been saved as：{excel_path}")

    def yolo_to_bbox(self, x_center, y_center, w, h, img_width, img_height, padding=0.1):
        # 计算原始边界框的像素坐标
        x_min = int((x_center - w / 2) * img_width)
        y_min = int((y_center - h / 2) * img_height)
        x_max = int((x_center + w / 2) * img_width)
        y_max = int((y_center + h / 2) * img_height)

        # 计算padding的像素值
        pad_x = int(padding * img_width)
        pad_y = int(padding * img_height)

        # 扩展边界框
        x_min = max(x_min - pad_x, 0)
        y_min = max(y_min - pad_y, 0)
        x_max = min(x_max + pad_x, img_width)
        y_max = min(y_max + pad_y, img_height)

        return x_min, y_min, x_max, y_max
        
    def crop_locks(self, image_path, label_path, output_folder, padding=0.1):
        # 读取原始图像
        img = cv2.imread(image_path)
        if img is None:
            print(f"Cannot read: {image_path}")
            return None

        img_height, img_width = img.shape[:2]

        # 读取对应的标签文件
        label_file = os.path.splitext(os.path.basename(image_path))[0] + '.txt'
        label_path = os.path.join(label_path, label_file)
        if not os.path.exists(label_path):
            print(f"label txt doesn't exist: {label_path}")
            return None

        with open(label_path, 'r') as f:
            lines = f.readlines()

        # 遍历每个检测到的目标
        for i, line in enumerate(lines):
            # 解析YOLO格式的标签
            class_id, x_center, y_center, w, h = map(float, line.strip().split())
            # 锁的类别ID为0和1
            if int(class_id) == 0 or int(class_id) == 1:
                # 转换为像素坐标，并添加padding
                x_min, y_min, x_max, y_max = self.yolo_to_bbox(x_center, y_center, w, h, img_width, img_height, padding)
                # 裁剪图像
                lock_img = img[y_min:y_max, x_min:x_max]
                if lock_img.size == 0:
                    print(f"裁剪区域无效: {image_path} 中的锁 {i+1}")
                    continue
                # 保存裁剪后的图像
                output_filename = f"{os.path.splitext(os.path.basename(image_path))[0]}.bmp"
                output_path = os.path.join(output_folder, output_filename)
                cv2.imwrite(output_path, lock_img)
                print(f"Save lock cut to: {output_path}")
                return lock_img  # 返回裁剪后的图片用于显示
        return None

    def process_and_classify_image(self, image_path):
        # 设置文件夹路径
        intermediate_folder1 = os.path.join(self.predict_folder, "intermediate1")
        intermediate_folder2 = os.path.join(self.predict_folder, "intermediate2")
        label_folder = os.path.join(intermediate_folder1, "labels")

        # 确保文件夹存在
        os.makedirs(intermediate_folder1, exist_ok=True)
        os.makedirs(intermediate_folder2, exist_ok=True)
        os.makedirs(label_folder, exist_ok=True)
        os.makedirs(self.inside_folder, exist_ok=True)
        os.makedirs(self.outside_folder, exist_ok=True)

        # 1. 使用model3检测锁的位置
        results3 = self.model3(image_path, save_txt=True, project=self.predict_folder, name="intermediate1", exist_ok=True)

        # 2. 裁剪锁图片
        cropped_img = self.crop_locks(image_path, label_folder, intermediate_folder2, padding=0.015)
        
        # 获取裁剪后的图片路径
        base_name = os.path.splitext(os.path.basename(image_path))[0]
        cropped_path = os.path.join(intermediate_folder2, base_name + '.bmp')

        # 3. 检测锁内部缺陷
        results1 = self.model1(cropped_path, save=True, save_txt=True, project=self.predict_folder, name="Inside", exist_ok=True)

        # 4. 检测锁外部缺陷
        results2 = self.model2(image_path, save=True, save_txt=True, project=self.predict_folder, name="Outside", exist_ok=True)

        # 检查是否有缺陷
        has_defects = False
        for result in [results1, results2]:
            if len(result[0].boxes) > 0:
                has_defects = True
                break

        # 生成Excel文件
        self.generate_excel()

        # 清理临时文件夹
        shutil.rmtree(intermediate_folder1)
        shutil.rmtree(intermediate_folder2)

        # 获取处理后的图片路径
        inside_path = os.path.join(self.inside_folder, base_name + '.jpg')
        outside_path = os.path.join(self.outside_folder, base_name + '.jpg')
        processed_image_path = inside_path if os.path.exists(inside_path) else outside_path

        # 读取处理后的图片
        processed_img = cv2.imread(processed_image_path) if os.path.exists(processed_image_path) else cv2.imread(image_path)

        # 返回结果和处理后的图片
        return "NG" if has_defects else "OK", processed_img

class FileMonitorThread(QThread):
    image_processed = pyqtSignal(str, str, QImage)
    stats_updated = pyqtSignal(dict)

    class FileHandler(FileSystemEventHandler):
        def __init__(self, callback):
            self.callback = callback
            super().__init__()

        def on_created(self, event):
            if event.is_directory:
                return
            if event.src_path.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png')):
                self.callback(event.src_path)

    def __init__(self, watch_folder):
        super().__init__()
        self.watch_folder = watch_folder
        self.running = False
        self.processor = ImageProcessor(watch_folder)
        
        # 设置所有需要的文件夹路径
        self.ok_folder = os.path.join(self.watch_folder, "OK")
        self.ng_folder = os.path.join(self.watch_folder, "NG")
        
        # 确保必要的文件夹存在
        os.makedirs(self.ok_folder, exist_ok=True)
        os.makedirs(self.ng_folder, exist_ok=True)

    def update_stats(self):
        stats = {
            'total': len(os.listdir(self.ok_folder)) + len(os.listdir(self.ng_folder)),
            'ok_count': len([f for f in os.listdir(self.ok_folder) if not os.path.isdir(os.path.join(self.ok_folder, f))]),
            'ng_count': len([f for f in os.listdir(self.ng_folder) if not os.path.isdir(os.path.join(self.ng_folder, f))]),
            'pending': len([f for f in os.listdir(self.watch_folder) 
                          if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))
                          and os.path.isfile(os.path.join(self.watch_folder, f))])
        }
        if stats['total'] > 0:
            stats['ok_rate'] = (stats['ok_count'] / stats['total']) * 100
        else:
            stats['ok_rate'] = 0
        self.stats_updated.emit(stats)

    def process_image(self, image_path):
        try:
            # 读取原始图片用于显示
            original_img = cv2.imread(image_path)
            if original_img is None:
                print(f"无法读取图片: {image_path}")
                return

            # 转换原始图片格式用于显示
            rgb_img = cv2.cvtColor(original_img, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_img.shape
            bytes_per_line = ch * w
            qt_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888)

            # 处理图片
            result, _ = self.processor.process_and_classify_image(image_path)
            
            # 移动原图到OK/NG文件夹
            filename = os.path.basename(image_path)
            
            if result == "OK":
                dest_path = os.path.join(self.ok_folder, filename)
            else:
                dest_path = os.path.join(self.ng_folder, filename)
            
            # 复制原图到对应文件夹
            shutil.copy2(image_path, dest_path)
            os.remove(image_path)  # 删除原图
            
            # 发送信号更新UI
            self.image_processed.emit(dest_path, result, qt_img)
            self.update_stats()
            
        except Exception as e:
            print(f"处理图片时出错 {image_path}: {str(e)}")

    def run(self):
        self.running = True
        event_handler = self.FileHandler(self.process_image)
        observer = Observer()
        observer.schedule(event_handler, self.watch_folder, recursive=False)
        observer.start()
        
        # 处理现有文件
        existing_files = [f for f in os.listdir(self.watch_folder)
                        if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))
                        and os.path.isfile(os.path.join(self.watch_folder, f))]
        for filename in existing_files:
            self.process_image(os.path.join(self.watch_folder, filename))
        
        # 保持线程运行
        while self.running:
            time.sleep(1)
        
        observer.stop()
        observer.join()

    def stop(self):
        self.running = False
        # 等待所有处理完成
        self.wait()
        # 更新最终统计
        self.update_stats()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("缺陷检测系统")
        self.setGeometry(100, 100, 1200, 900)  # 增加窗口大小
        
        # 添加工作目录变量
        self.work_dir = None
        self.predict_dir = None
        self.watch_folder = None  # 初始化为None
        
        # 设置整体样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
                color: #757575;
            }
            QPushButton:disabled QIcon {
                opacity: 0.5;
            }
            QLabel {
                font-size: 14px;
            }
            QGroupBox {
                border: 2px solid #E0E0E0;
                border-radius: 6px;
                margin-top: 12px;
                font-size: 14px;
                padding: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QTableWidget {
                border: 1px solid #E0E0E0;
                gridline-color: #E0E0E0;
                background-color: white;
                alternate-background-color: #FAFAFA;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 5px;
                border: 1px solid #E0E0E0;
                font-weight: bold;
            }
        """)
        
        # 创建主窗口部件
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setSpacing(20)  # 增加组件之间的间距
        layout.setContentsMargins(20, 20, 20, 20)  # 设置边距
        
        # 创建控制面板组
        control_group = QGroupBox("控制面板")
        button_layout = QHBoxLayout()
        
        # 创建左侧按钮布局
        left_button_layout = QHBoxLayout()
        
        # 创建选择文件夹按钮
        self.select_folder_button = QPushButton("选择文件夹")
        self.select_folder_button.setIcon(QIcon("icons/文件夹.svg"))
        self.select_folder_button.setIconSize(QSize(32, 32))
        
        # 创建其他按钮
        self.start_button = QPushButton("运行")
        self.start_button.setIcon(QIcon("icons/运行.svg"))
        self.start_button.setIconSize(QSize(37, 37))
        
        self.stop_button = QPushButton("停止")
        self.stop_button.setIcon(QIcon("icons/结束.svg"))
        self.stop_button.setIconSize(QSize(32, 32))
        
        self.view_results_button = QPushButton("查看结果")
        self.view_results_button.setIcon(QIcon("icons/查看结果.svg"))
        self.view_results_button.setIconSize(QSize(32, 32))
        
        self.reset_data_button = QPushButton("重置数据")
        self.reset_data_button.setIcon(QIcon("icons/重置数据.svg"))
        self.reset_data_button.setIconSize(QSize(32, 32))
        
        # 设置按钮样式
        for button in [self.select_folder_button, self.start_button, self.stop_button, 
                      self.view_results_button, self.reset_data_button]:
            button.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    color: #333333;
                    border: none;
                    border-radius: 4px;
                    font-size: 16px;
                    min-width: 120px;
                    min-height: 45px;
                    text-align: center;
                    padding: 0px;
                }
                QPushButton:hover {
                    color: #666666;
                }
                QPushButton:disabled {
                    color: #BDBDBD;
                }
            """)
            button.setLayoutDirection(Qt.LeftToRight)
            button.setToolButtonStyle = Qt.ToolButtonTextBesideIcon
        
        # 连接按钮信号
        self.select_folder_button.clicked.connect(self.select_work_folder)
        self.start_button.clicked.connect(self.start_monitoring)
        self.stop_button.clicked.connect(self.stop_monitoring)
        self.view_results_button.clicked.connect(self.view_results)
        self.reset_data_button.clicked.connect(self.reset_data)
        
        # 初始状态设置
        self.start_button.setEnabled(False)  # 初始禁用运行按钮
        self.stop_button.setEnabled(False)
        
        # 添加按钮到左侧布局
        left_button_layout.addWidget(self.select_folder_button)
        left_button_layout.addWidget(self.start_button)
        left_button_layout.addWidget(self.stop_button)
        left_button_layout.addWidget(self.view_results_button)
        left_button_layout.addWidget(self.reset_data_button)
        
        # 创建右侧时间信息布局
        right_time_layout = QHBoxLayout()
        
        # 创建时间标签
        self.start_time_label = QLabel("开启时间: --")
        self.run_time_label = QLabel("运行时间: 00:00:00")
        
        # 创建待机时间标签
        self.idle_time_label = QLabel("待机时间: 00:00:00")
        self.idle_time_label.setStyleSheet("""
            QLabel {
                color: #333333;
                font-size: 14px;
                padding: 5px 10px;
                background-color: #f8f8f8;
                border-radius: 4px;
                margin: 5px;
            }
        """)
        
        # 设置时间标签样式
        time_label_style = """
            QLabel {
                color: #333333;
                font-size: 14px;
                padding: 5px 10px;
                background-color: #f8f8f8;
                border-radius: 4px;
                margin: 5px;
            }
        """
        self.start_time_label.setStyleSheet(time_label_style)
        self.run_time_label.setStyleSheet(time_label_style)
        
        # 添加时间标签到右侧布局
        right_time_layout.addWidget(self.start_time_label)
        right_time_layout.addWidget(self.run_time_label)
        right_time_layout.addWidget(self.idle_time_label)
        
        # 添加待机时间计时器
        self.idle_timer = QTimer()
        self.idle_timer.timeout.connect(self.update_idle_time)
        self.idle_timer.setInterval(1000)  # 每秒更新一次
        self.idle_start_time = None
        
        # 将左侧按钮布局和右侧时间布局添加到主布局
        button_layout.addLayout(left_button_layout)
        button_layout.addStretch()  # 添加弹性空间
        button_layout.addLayout(right_time_layout)
        
        control_group.setLayout(button_layout)
        layout.addWidget(control_group)
        
        # 创建图片显示组
        image_group = QGroupBox("图片预览")
        image_layout = QVBoxLayout()
        
        # 添加图片名称标签
        self.image_name_label = QLabel()
        self.image_name_label.setAlignment(Qt.AlignCenter)
        self.image_name_label.setStyleSheet("""
            QLabel {
                color: #666666;
                font-size: 12pt;
                padding: 5px;
                background-color: #f8f8f8;
                border-radius: 4px;
                margin-bottom: 5px;
            }
        """)
        image_layout.addWidget(self.image_name_label)
        
        # 创建水平布局用于放置图片和导航按钮
        image_container = QHBoxLayout()
        
        # 创建上一张和下一张按钮
        self.prev_image_button = QPushButton()
        self.next_image_button = QPushButton()
        
        # 设置按钮图标
        self.prev_image_button.setIcon(QIcon("icons/图片切换-左.svg"))
        self.next_image_button.setIcon(QIcon("icons/图片切换-右.svg"))
        
        # 设置图标大小
        icon_size = QSize(20, 20)  # 减小图标大小
        self.prev_image_button.setIconSize(icon_size)
        self.next_image_button.setIconSize(icon_size)
        
        # 设置按钮大小
        button_size = QSize(24, 60)  # 减小按钮大小
        self.prev_image_button.setFixedSize(button_size)
        self.next_image_button.setFixedSize(button_size)
        
        # 设置导航按钮的基本样式
        nav_button_style = """
            QPushButton {
                background-color: transparent;
                border: none;
                padding: 0px;
                margin: 0px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border-radius: 4px;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
                border-radius: 4px;
            }
            QPushButton:disabled {
                background-color: transparent;
                qproperty-iconColor: rgba(189, 189, 189, 0.4);
            }
            QPushButton:enabled {
                background-color: transparent;
                qproperty-iconColor: #000000;
            }
        """
        self.prev_image_button.setStyleSheet(nav_button_style)
        self.next_image_button.setStyleSheet(nav_button_style)
        
        # 初始状态下禁用切换按钮
        self.prev_image_button.setEnabled(False)
        self.next_image_button.setEnabled(False)
        
        # 连接按钮信号
        self.prev_image_button.clicked.connect(self.show_previous_image)
        self.next_image_button.clicked.connect(self.show_next_image)
        
        # 创建图片标签
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setMinimumSize(600, 400)
        self.image_label.setStyleSheet("""
            QLabel {
                background-color: white;
                border: 2px solid #E0E0E0;
                border-radius: 4px;
            }
        """)
        
        # 添加一些间距
        image_container.addSpacing(10)
        
        # 将按钮和图片添加到水平布局中
        image_container.addWidget(self.prev_image_button)
        image_container.addWidget(self.image_label)
        image_container.addWidget(self.next_image_button)
        
        # 添加一些间距
        image_container.addSpacing(10)
        
        # 将image_container添加到image_layout
        image_layout.addLayout(image_container)
        
        # 创建缺陷查看按钮布局
        defect_button_layout = QHBoxLayout()
        defect_button_layout.setContentsMargins(0, 0, 0, 0)  # 减少边距
        
        # 创建查看缺陷按钮
        self.view_inside_button = QPushButton("查看锁内缺陷")
        self.view_outside_button = QPushButton("查看锁外缺陷")
        
        # 设置缺陷查看按钮样式
        defect_button_style = """
            QPushButton {
                background-color: #f5f5f5;
                color: #333333;
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                font-size: 16px;
                font-weight: bold;
                min-width: 150px;
                min-height: 30px;
                padding: 4px 15px;
                margin: 5px 2px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border-color: #BDBDBD;
            }
            QPushButton:disabled {
                color: #BDBDBD;
                background-color: #f9f9f9;
                border-color: #E0E0E0;
            }
        """
        self.view_inside_button.setStyleSheet(defect_button_style)
        self.view_outside_button.setStyleSheet(defect_button_style)
        
        # 初始状态下禁用缺陷查看按钮
        self.view_inside_button.setEnabled(False)
        self.view_outside_button.setEnabled(False)
        
        # 连接按钮信号
        self.view_inside_button.clicked.connect(self.view_inside_defect)
        self.view_outside_button.clicked.connect(self.view_outside_defect)
        
        # 添加缺陷查看按钮到布局
        defect_button_layout.addWidget(self.view_inside_button)
        defect_button_layout.addWidget(self.view_outside_button)
        
        image_layout.addLayout(defect_button_layout)
        
        self.result_label = QLabel()
        self.result_label.setAlignment(Qt.AlignCenter)
        self.result_label.setStyleSheet("QLabel { font-size: 18pt; font-weight: bold; }")
        
        image_layout.addWidget(self.result_label)
        image_group.setLayout(image_layout)
        layout.addWidget(image_group)
        
        # 创建统计信息组
        stats_group = QGroupBox("实时统计")
        stats_layout = QHBoxLayout()
        
        self.total_label = QLabel("已处理总数: 0")
        self.ok_count_label = QLabel("良品数量: 0")
        self.ok_rate_label = QLabel("良品率: 0%")
        self.ng_count_label = QLabel("不良品数量: 0")
        self.pending_label = QLabel("待处理数量: 0")
        
        for label in [self.total_label, self.ok_count_label, self.ok_rate_label, 
                     self.ng_count_label, self.pending_label]:
            label.setStyleSheet("""
                QLabel {
                    padding: 8px 15px;
                    background-color: white;
                    border: 1px solid #E0E0E0;
                    border-radius: 4px;
                    font-size: 10pt;
                    font-weight: bold;
                    color: #333333;
                }
            """)
            
        stats_layout.addWidget(self.total_label)
        stats_layout.addWidget(self.ok_count_label)
        stats_layout.addWidget(self.ok_rate_label)
        stats_layout.addWidget(self.ng_count_label)
        stats_layout.addWidget(self.pending_label)
        
        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        # 创建缺陷统计表格组
        table_group = QGroupBox("缺陷统计")
        table_layout = QVBoxLayout()
        
        self.defect_table = QTableWidget()
        self.defect_table.setColumnCount(10)  # 修改为10列
        self.defect_table.setHorizontalHeaderLabels([
            "文件夹", "总数", "良品数", "良品率", "不良总数", 
            "毛边", "针孔", "欠缺", "划伤", "油污"
        ])
        self.defect_table.setRowCount(1)
        self.defect_table.setAlternatingRowColors(True)  # 设置交替行颜色
        
        # 设置表格的样式
        self.defect_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.defect_table.setMinimumHeight(100)
        self.defect_table.setMaximumHeight(150)
        
        # 初始化表格内容
        for col in range(10):  # 修改为10列
            item = QTableWidgetItem("0")
            item.setTextAlignment(Qt.AlignCenter)
            self.defect_table.setItem(0, 0, QTableWidgetItem("Picture"))
            self.defect_table.setItem(0, col, item)
        
        table_layout.addWidget(self.defect_table)
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)
        
        # 初始化监控线程
        self.monitor_thread = None
        self.watch_folder = None  # 移除固定值设置
        
        # 初始化时间相关变量
        self.start_datetime = None
        self.run_timer = QTimer()
        self.run_timer.timeout.connect(self.update_run_time)
        self.run_timer.setInterval(1000)  # 每秒更新一次
        
        # 在__init__方法中添加新的成员变量
        self.processed_images = []  # 存储已处理图片的路径
        self.current_image_index = -1  # 当前显示的图片索引
        
    def select_work_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择工作目录")
        if folder:
            self.work_dir = folder
            self.watch_folder = folder  # 更新watch_folder为选择的工作目录
            # 设置预测结果目录
            self.predict_dir = os.path.join(folder, "Predict")
            
            # 创建必要的文件夹结构
            os.makedirs(os.path.join(folder, "OK"), exist_ok=True)
            os.makedirs(os.path.join(folder, "NG"), exist_ok=True)
            os.makedirs(self.predict_dir, exist_ok=True)
            
            # 更新UI状态
            self.start_button.setEnabled(True)
            self.image_label.clear()
            self.image_name_label.clear()
            self.result_label.clear()
            
            # 重置统计数据
            stats = {
                'total': 0,
                'ok_count': 0,
                'ng_count': 0,
                'pending': 0,
                'ok_rate': 0
            }
            self.update_stats(stats)

    def start_monitoring(self):
        if self.work_dir:  # 确保已选择工作目录
            if not self.monitor_thread:
                self.monitor_thread = FileMonitorThread(self.work_dir)
                self.monitor_thread.image_processed.connect(self.update_image)
                self.monitor_thread.stats_updated.connect(self.update_stats)
                self.monitor_thread.start()
                self.start_button.setEnabled(False)
                self.stop_button.setEnabled(True)
                self.select_folder_button.setEnabled(False)  # 禁用文件夹选择按钮
                
                # 设置导航按钮为透明但保持位置
                transparent_style = """
                    QPushButton {
                        background-color: transparent;
                        border: none;
                        padding: 0px;
                        margin: 0px;
                        opacity: 0;
                    }
                    QPushButton:hover {
                        background-color: transparent;
                    }
                    QPushButton:pressed {
                        background-color: transparent;
                    }
                    QPushButton:disabled {
                        background-color: transparent;
                    }
                    QPushButton:enabled {
                        opacity: 0;
                    }
                """
                self.prev_image_button.setStyleSheet(transparent_style)
                self.next_image_button.setStyleSheet(transparent_style)
                
                # 更新开启时间和启动运行时间计时器
                self.start_datetime = datetime.now()
                self.start_time_label.setText(f"开启时间: {self.start_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
                self.run_timer.start()
                
                # 停止待机时间计时器并重置
                self.idle_timer.stop()
                self.idle_start_time = None
                self.idle_time_label.setText("待机时间: 00:00:00")

    def stop_monitoring(self):
        if self.monitor_thread:
            # 停止监控线程
            self.monitor_thread.stop()
            # 等待线程完全结束
            self.monitor_thread.wait()
            self.monitor_thread = None
            
            # 更新UI状态
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)
            self.select_folder_button.setEnabled(True)  # 启用文件夹选择按钮
            
            # 恢复导航按钮的完全可见性
            nav_button_style = """
                QPushButton {
                    background-color: transparent;
                    border: none;
                    padding: 0px;
                    margin: 0px;
                }
                QPushButton:hover {
                    background-color: #e0e0e0;
                    border-radius: 4px;
                }
                QPushButton:pressed {
                    background-color: #d0d0d0;
                    border-radius: 4px;
                }
                QPushButton:disabled {
                    background-color: transparent;
                    qproperty-iconColor: rgba(189, 189, 189, 0.4);
                }
                QPushButton:enabled {
                    background-color: transparent;
                    qproperty-iconColor: #000000;
                }
            """
            self.prev_image_button.setStyleSheet(nav_button_style)
            self.next_image_button.setStyleSheet(nav_button_style)
            
            # 停止运行时间计时器
            self.run_timer.stop()
            
            # 启动待机时间计时器
            self.idle_start_time = datetime.now()
            self.idle_timer.start()
            
            # 启用导航按钮并更新状态
            self.update_nav_buttons_state(load_first_image=True)
            
            # 启用缺陷查看按钮
            self.view_inside_button.setEnabled(True)
            self.view_outside_button.setEnabled(True)

    def view_results(self):
        if self.work_dir:
            # 打开Excel结果文件
            excel_path = os.path.join(self.work_dir, "Predict.xlsx")
            try:
                if os.path.exists(excel_path):
                    os.startfile(excel_path)
                else:
                    from PyQt5.QtWidgets import QMessageBox
                    QMessageBox.information(self, "提示", "结果文件不存在")
            except Exception as e:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.warning(self, "错误", f"打开结果文件时出错: {str(e)}")
    
    def update_image(self, image_path, result, qt_img):
        # 更新图片显示
        if qt_img is not None:
            scaled_pixmap = QPixmap.fromImage(qt_img).scaled(
                self.image_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.image_label.setPixmap(scaled_pixmap)
            
            # 更新图片名称
            self.image_name_label.setText(f"当前图片: {os.path.basename(image_path)}")
            
            # 根据结果设置不同的样式
            if result == "OK":
                color = "green"
            else:
                color = "red"
            
            # 更新结果标签
            self.result_label.setStyleSheet(f"QLabel {{ color: {color}; font-size: 18pt; font-weight: bold; }}")
            self.result_label.setText(f"检测结果: {result}")
            
            # 在监控运行时设置导航按钮为透明
            nav_button_style = """
                QPushButton {
                    background-color: transparent;
                    border: none;
                    padding: 0px;
                    margin: 0px;
                    opacity: 0;
                }
                QPushButton:hover {
                    background-color: transparent;
                }
                QPushButton:pressed {
                    background-color: transparent;
                }
                QPushButton:disabled {
                    background-color: transparent;
                }
            """
            self.prev_image_button.setStyleSheet(nav_button_style)
            self.next_image_button.setStyleSheet(nav_button_style)
            self.view_inside_button.setEnabled(False)
            self.view_outside_button.setEnabled(False)
    
    def update_stats(self, stats):
        self.total_label.setText(f"已处理总数: {stats['total']}")
        self.ok_count_label.setText(f"良品数量: {stats['ok_count']}")
        self.ok_rate_label.setText(f"良品率: {stats['ok_rate']:.2f}%")
        self.ng_count_label.setText(f"不良品数量: {stats['ng_count']}")
        self.pending_label.setText(f"待处理数量: {stats['pending']}")
        # 更新缺陷统计表格
        self.update_defect_stats()

    def get_class_name(self, label_folder, label_id):
        classes_file = os.path.join(label_folder, "classes.txt")
        if not os.path.exists(classes_file):
            return "Unknown"
        with open(classes_file, "r") as f:
            classes = [line.strip() for line in f.readlines()]
        return classes[int(label_id)] if int(label_id) < len(classes) else "Unknown"

    def update_defect_stats(self):
        try:
            # 获取Picture文件夹下的OK和NG文件夹
            ok_folder = os.path.join(self.watch_folder, "OK")
            ng_folder = os.path.join(self.watch_folder, "NG")
            
            # 初始化计数器
            total_count = 0
            ok_count = 0
            defect_counts = {
                "maobian": set(),
                "zhenkong": set(),
                "qianque": set(),
                "huashang": set(),
                "youwu": set()
            }

            # 统计OK文件夹中的图片
            if os.path.exists(ok_folder):
                ok_files = [f for f in os.listdir(ok_folder) if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
                ok_count = len(ok_files)
                total_count += ok_count

            # 统计NG文件夹中的图片和缺陷类型
            if os.path.exists(ng_folder):
                ng_files = [f for f in os.listdir(ng_folder) if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
                total_count += len(ng_files)
                
                # 对于每个NG图片，查找对应的标签文件来统计缺陷类型
                for ng_file in ng_files:
                    base_name = os.path.splitext(ng_file)[0]
                    
                    # 检查Inside和Outside文件夹中的标签
                    inside_label = os.path.join(self.work_dir, "Predict/Inside/labels", f"{base_name}.txt")
                    outside_label = os.path.join(self.work_dir, "Predict/Outside/labels", f"{base_name}.txt")
                    
                    # 记录该图片的所有缺陷类型
                    for label_file in [inside_label, outside_label]:
                        if os.path.exists(label_file):
                            with open(label_file, 'r') as f:
                                lines = f.readlines()
                                for line in lines:
                                    class_id = int(line.split()[0])
                                    class_name = self.get_class_name(os.path.dirname(label_file), class_id)
                                    if class_name in defect_counts:
                                        defect_counts[class_name].add(base_name)

            # 计算良品率和不良总数
            ok_rate = (ok_count / total_count * 100) if total_count > 0 else 0
            ng_total = total_count - ok_count

            # 更新表格
            self.defect_table.setItem(0, 0, QTableWidgetItem(os.path.basename(self.watch_folder)))
            self.defect_table.setItem(0, 1, QTableWidgetItem(str(total_count)))
            self.defect_table.setItem(0, 2, QTableWidgetItem(str(ok_count)))
            self.defect_table.setItem(0, 3, QTableWidgetItem(f"{ok_rate:.2f}%"))
            self.defect_table.setItem(0, 4, QTableWidgetItem(str(ng_total)))
            self.defect_table.setItem(0, 5, QTableWidgetItem(str(len(defect_counts["maobian"]))))
            self.defect_table.setItem(0, 6, QTableWidgetItem(str(len(defect_counts["zhenkong"]))))
            self.defect_table.setItem(0, 7, QTableWidgetItem(str(len(defect_counts["qianque"]))))
            self.defect_table.setItem(0, 8, QTableWidgetItem(str(len(defect_counts["huashang"]))))
            self.defect_table.setItem(0, 9, QTableWidgetItem(str(len(defect_counts["youwu"]))))
            
            # 设置所有单元格居中对齐
            for col in range(10):  # 修改为10列
                item = self.defect_table.item(0, col)
                if item:
                    item.setTextAlignment(Qt.AlignCenter)
                    
        except Exception as e:
            print(f"更新缺陷统计时出错: {str(e)}")

    def update_run_time(self):
        if self.start_datetime:
            elapsed = datetime.now() - self.start_datetime
            hours = int(elapsed.total_seconds() // 3600)
            minutes = int((elapsed.total_seconds() % 3600) // 60)
            seconds = int(elapsed.total_seconds() % 60)
            self.run_time_label.setText(f"运行时间: {hours:02d}:{minutes:02d}:{seconds:02d}")

    def update_idle_time(self):
        if self.idle_start_time:
            elapsed = datetime.now() - self.idle_start_time
            hours = int(elapsed.total_seconds() // 3600)
            minutes = int((elapsed.total_seconds() % 3600) // 60)
            seconds = int(elapsed.total_seconds() % 60)
            self.idle_time_label.setText(f"待机时间: {hours:02d}:{minutes:02d}:{seconds:02d}")

    def reset_data(self):
        try:
            # 如果监控正在运行，先停止
            if self.monitor_thread:
                self.stop_monitoring()
            
            if self.work_dir:  # 确保有工作目录
                # 重置时间显示
                self.start_datetime = None
                self.start_time_label.setText("开启时间: --")
                self.run_time_label.setText("运行时间: 00:00:00")
                self.idle_start_time = None
                self.idle_timer.stop()
                self.idle_time_label.setText("待机时间: 00:00:00")
                
                # 重置图片切换相关变量
                self.processed_images = []
                self.current_image_index = -1
                
                # 隐藏导航按钮
                self.prev_image_button.hide()
                self.next_image_button.hide()
                self.update_nav_buttons_state()
                
                # 创建以当前时间命名的目标文件夹（在代码运行根目录下）
                current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                result_base_dir = "./Result"  # 修改为代码运行根目录
                target_dir = os.path.join(result_base_dir, current_time)
                
                # 确保Result文件夹存在
                os.makedirs(result_base_dir, exist_ok=True)
                os.makedirs(target_dir, exist_ok=True)
                
                # 移动Predict文件夹
                predict_dir = os.path.join(self.work_dir, "Predict")
                if os.path.exists(predict_dir):
                    target_predict = os.path.join(target_dir, "Predict")
                    shutil.move(predict_dir, target_predict)
                
                # 移动Predict.xlsx
                excel_file = os.path.join(self.work_dir, "Predict.xlsx")
                if os.path.exists(excel_file):
                    target_excel = os.path.join(target_dir, "Predict.xlsx")
                    shutil.move(excel_file, target_excel)
                
                # 处理OK和NG文件夹
                ok_dir = os.path.join(self.work_dir, "OK")
                ng_dir = os.path.join(self.work_dir, "NG")
                
                # 创建目标文件夹
                target_ok = os.path.join(target_dir, "OK")
                target_ng = os.path.join(target_dir, "NG")
                os.makedirs(target_ok, exist_ok=True)
                os.makedirs(target_ng, exist_ok=True)
                
                # 移动OK文件夹中的文件并删除文件夹
                if os.path.exists(ok_dir):
                    for file in os.listdir(ok_dir):
                        src_file = os.path.join(ok_dir, file)
                        dst_file = os.path.join(target_ok, file)
                        if os.path.isfile(src_file):
                            shutil.move(src_file, dst_file)
                    # 删除原OK文件夹
                    shutil.rmtree(ok_dir)
                
                # 移动NG文件夹中的文件并删除文件夹
                if os.path.exists(ng_dir):
                    for file in os.listdir(ng_dir):
                        src_file = os.path.join(ng_dir, file)
                        dst_file = os.path.join(target_ng, file)
                        if os.path.isfile(src_file):
                            shutil.move(src_file, dst_file)
                    # 删除原NG文件夹
                    shutil.rmtree(ng_dir)
                
                # 重新创建空的OK和NG文件夹
                os.makedirs(ok_dir, exist_ok=True)
                os.makedirs(ng_dir, exist_ok=True)
                
                # 清空图片预览和统计信息
                self.image_label.clear()
                self.image_name_label.clear()
                self.result_label.clear()
                
                # 重置统计数据
                stats = {
                    'total': 0,
                    'ok_count': 0,
                    'ng_count': 0,
                    'pending': 0,
                    'ok_rate': 0
                }
                self.update_stats(stats)
                
                # 删除工作目录中的OK和NG文件夹
                try:
                    ok_dir = os.path.join(self.work_dir, "OK")
                    ng_dir = os.path.join(self.work_dir, "NG")
                    if os.path.exists(ok_dir):
                        shutil.rmtree(ok_dir)
                    if os.path.exists(ng_dir):
                        shutil.rmtree(ng_dir)
                except Exception as e:
                    print(f"删除OK/NG文件夹时出错: {str(e)}")
                
                # 提示用户重置完成
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.information(self, "重置完成", f"所有数据已移动到Result/{current_time}文件夹中")
                
        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "重置错误", f"重置数据时出错：{str(e)}")

    def update_nav_buttons_state(self, load_first_image=True):
        # 检查是否已设置工作目录
        if not self.watch_folder:
            return
            
        # 获取所有已处理的图片
        ok_folder = os.path.join(self.watch_folder, "OK")
        ng_folder = os.path.join(self.watch_folder, "NG")
        
        all_images = []
        
        # 收集OK文件夹中的图片
        if os.path.exists(ok_folder):
            ok_files = [os.path.join(ok_folder, f) for f in os.listdir(ok_folder) 
                      if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
            all_images.extend(ok_files)
        
        # 收集NG文件夹中的图片
        if os.path.exists(ng_folder):
            ng_files = [os.path.join(ng_folder, f) for f in os.listdir(ng_folder)
                      if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
            all_images.extend(ng_files)
        
        # 按文件名排序
        all_images.sort(key=lambda x: os.path.basename(x))
        
        # 如果监控线程正在运行，禁用所有导航按钮
        if self.monitor_thread is not None:
            self.prev_image_button.setEnabled(False)
            self.next_image_button.setEnabled(False)
            self.view_inside_button.setEnabled(False)
            self.view_outside_button.setEnabled(False)
            return
        
        # 设置导航按钮的基本样式
        nav_button_style = """
            QPushButton {
                background-color: transparent;
                border: none;
                padding: 0px;
                margin: 0px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border-radius: 4px;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
                border-radius: 4px;
            }
            QPushButton:disabled {
                background-color: transparent;
                qproperty-iconColor: rgba(189, 189, 189, 0.4);
            }
            QPushButton:enabled {
                background-color: transparent;
                qproperty-iconColor: #000000;
            }
        """
        self.prev_image_button.setStyleSheet(nav_button_style)
        self.next_image_button.setStyleSheet(nav_button_style)
        
        # 找到当前显示的图片在列表中的位置
        current_image = None
        if self.image_label.pixmap():
            current_name = self.image_name_label.text().replace("当前图片: ", "")
            for i, img_path in enumerate(all_images):
                if os.path.basename(img_path) == current_name:
                    current_image = i
                    break
        
        # 如果没有找到当前图片，但有图片可显示，则显示第一张
        if current_image is None and all_images and load_first_image:
            current_image = 0
            self.load_image_at_index(all_images, current_image, update_buttons=False)
        
        # 更新按钮状态
        has_images = len(all_images) > 0
        if current_image is not None:
            self.prev_image_button.setEnabled(current_image > 0)
            self.next_image_button.setEnabled(current_image < len(all_images) - 1)
            # 启用缺陷查看按钮
            self.view_inside_button.setEnabled(True)
            self.view_outside_button.setEnabled(True)
        else:
            self.prev_image_button.setEnabled(False)
            self.next_image_button.setEnabled(False)
            self.view_inside_button.setEnabled(False)
            self.view_outside_button.setEnabled(False)

    def load_image_at_index(self, all_images, index, update_buttons=True):
        if 0 <= index < len(all_images):
            image_path = all_images[index]
            # 读取图片
            img = cv2.imread(image_path)
            if img is not None:
                # 转换为RGB格式
                rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                h, w, ch = rgb_img.shape
                bytes_per_line = ch * w
                qt_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888)
                
                # 更新图片显示
                scaled_pixmap = QPixmap.fromImage(qt_img).scaled(
                    self.image_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.image_label.setPixmap(scaled_pixmap)
                
                # 更新图片名称
                current_name = os.path.basename(image_path)
                self.image_name_label.setText(f"当前图片: {current_name}")
                
                # 判断图片结果（使用os.path来判断路径）
                result = "OK" if os.path.normpath(image_path).replace("\\", "/").find("/OK/") != -1 else "NG"
                color = "green" if result == "OK" else "red"
                
                # 更新检测结果显示
                self.result_label.setStyleSheet(f"QLabel {{ color: {color}; font-size: 18pt; font-weight: bold; }}")
                
                # 如果是NG，查找具体的缺陷类型
                if result == "NG":
                    base_name = os.path.splitext(current_name)[0]
                    defect_types = set()
                    
                    # 检查Inside和Outside文件夹中的标签
                    inside_label = os.path.join(self.work_dir, "Predict/Inside/labels", f"{base_name}.txt")
                    outside_label = os.path.join(self.work_dir, "Predict/Outside/labels", f"{base_name}.txt")
                    
                    for label_file in [inside_label, outside_label]:
                        if os.path.exists(label_file):
                            with open(label_file, 'r') as f:
                                lines = f.readlines()
                                for line in lines:
                                    class_id = int(line.split()[0])
                                    class_name = self.get_class_name(os.path.dirname(label_file), class_id)
                                    if class_name != "Unknown":
                                        defect_types.add(class_name)
                    
                    # 显示检测结果和缺陷类型
                    if defect_types:
                        defect_str = "、".join(defect_types)
                        self.result_label.setText(f"检测结果: {result} ({defect_str})")
                    else:
                        self.result_label.setText(f"检测结果: {result}")
                else:
                    self.result_label.setText(f"检测结果: {result}")
                
                # 更新导航按钮状态
                if update_buttons:
                    self.update_nav_buttons_state(load_first_image=False)
                    
                # 启用缺陷查看按钮
                self.view_inside_button.setEnabled(True)
                self.view_outside_button.setEnabled(True)

    def show_previous_image(self):
        ok_folder = os.path.join(self.watch_folder, "OK")
        ng_folder = os.path.join(self.watch_folder, "NG")
        
        all_images = []
        if os.path.exists(ok_folder):
            ok_files = [os.path.join(ok_folder, f) for f in os.listdir(ok_folder) 
                      if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
            all_images.extend(ok_files)
        
        if os.path.exists(ng_folder):
            ng_files = [os.path.join(ng_folder, f) for f in os.listdir(ng_folder)
                      if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
            all_images.extend(ng_files)
        
        all_images.sort(key=lambda x: os.path.basename(x))
        
        if not all_images:
            return
            
        current_name = self.image_name_label.text().replace("当前图片: ", "")
        current_index = -1
        
        # 如果当前没有显示图片，显示最后一张
        if not current_name:
            self.load_image_at_index(all_images, len(all_images) - 1)
            return
            
        # 查找当前图片的索引
        for i, img_path in enumerate(all_images):
            if os.path.basename(img_path) == current_name:
                current_index = i
                break
        
        # 如果找到当前图片并且不是第一张，显示上一张
        if current_index > 0:
            self.load_image_at_index(all_images, current_index - 1)
        # 如果没找到当前图片，显示最后一张
        elif current_index == -1:
            self.load_image_at_index(all_images, len(all_images) - 1)

    def show_next_image(self):
        ok_folder = os.path.join(self.watch_folder, "OK")
        ng_folder = os.path.join(self.watch_folder, "NG")
        
        all_images = []
        if os.path.exists(ok_folder):
            ok_files = [os.path.join(ok_folder, f) for f in os.listdir(ok_folder) 
                      if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
            all_images.extend(ok_files)
        
        if os.path.exists(ng_folder):
            ng_files = [os.path.join(ng_folder, f) for f in os.listdir(ng_folder)
                      if f.lower().endswith(('.bmp', '.jpg', '.jpeg', '.png'))]
            all_images.extend(ng_files)
        
        all_images.sort(key=lambda x: os.path.basename(x))
        
        if not all_images:
            return
            
        current_name = self.image_name_label.text().replace("当前图片: ", "")
        current_index = -1
        
        # 如果当前没有显示图片，显示第一张
        if not current_name:
            self.load_image_at_index(all_images, 0)
            return
            
        # 查找当前图片的索引
        for i, img_path in enumerate(all_images):
            if os.path.basename(img_path) == current_name:
                current_index = i
                break
        
        # 如果找到当前图片并且不是最后一张，显示下一张
        if current_index >= 0 and current_index < len(all_images) - 1:
            self.load_image_at_index(all_images, current_index + 1)
        # 如果没找到当前图片，显示第一张
        elif current_index == -1:
            self.load_image_at_index(all_images, 0)

    def view_inside_defect(self):
        if not self.image_name_label.text():
            return
            
        current_name = self.image_name_label.text().replace("当前图片: ", "")
        base_name = os.path.splitext(current_name)[0]
        inside_path = os.path.join(self.work_dir, "Predict/Inside", f"{base_name}.jpg")
        
        if os.path.exists(inside_path):
            try:
                os.startfile(inside_path)
            except Exception as e:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.warning(self, "打开失败", f"无法打开锁内缺陷图片：{str(e)}")
        else:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.information(self, "提示", "未找到对应的锁内缺陷图片")

    def view_outside_defect(self):
        if not self.image_name_label.text():
            return
            
        current_name = self.image_name_label.text().replace("当前图片: ", "")
        base_name = os.path.splitext(current_name)[0]
        outside_path = os.path.join(self.work_dir, "Predict/Outside", f"{base_name}.jpg")
        
        if os.path.exists(outside_path):
            try:
                os.startfile(outside_path)
            except Exception as e:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.warning(self, "打开失败", f"无法打开锁外缺陷图片：{str(e)}")
        else:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.information(self, "提示", "未找到对应的锁外缺陷图片")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_()) 